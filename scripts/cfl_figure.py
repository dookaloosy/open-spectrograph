"""Generate figures/fig_6_cfl_spectrum.png from a captured CFL spectrum.

Reads the spectrum workbook produced during calibration:
  Sheet1  — pixel / wavelength / counts, with the quadratic calibration
            constants a0, a1, a2 in cells B18:B20 (single source of truth)
  Sheet2  — Gaussian-fitted index-peak centers (px) and reference
            wavelengths (nm), one row per line

The main panel plots the full spectrum on the calibrated wavelength
axis and labels each index line.  The inset refits the isolated
Hg 404.7 nm line with a Gaussian and annotates its FWHM.

Usage:
    python3 scripts/cfl_figure.py [path/to/cfl_spectrum.xlsx]
"""

import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from scipy.optimize import curve_fit

REPO = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = (REPO / "output" / "export_czerny_baseline_v0_design"
                / "cfl_spectrum.xlsx")
OUT_PNG = REPO / "figures" / "fig_6_cfl_spectrum.png"

HG_LINE_NM = 404.656          # inset line (isolated, unsaturated)
INSET_XLIM = (401.0, 409.0)   # inset x-range (nm)
FIT_HALF_PX = 14              # Gaussian fit window half-width
# Main-axes and inset geometry measured off the original figure
# (figure fraction and main-axes fraction respectively).
MAIN_AXES_RECT = dict(left=0.0697, right=0.9803, bottom=0.1227, top=0.9659)
INSET_BOX = [0.720, 0.319, 0.260, 0.550]


def gauss(x, amp, mu, sigma, off):
    return amp * np.exp(-0.5 * ((x - mu) / sigma) ** 2) + off


def load_workbook_data(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    a0, a1, a2 = (float(ws[f"B{r}"].value) for r in (18, 19, 20))
    px, cnt = [], []
    for row in ws.iter_rows(values_only=True):
        if isinstance(row[0], (int, float)) and isinstance(row[2], (int, float)):
            px.append(row[0])
            cnt.append(row[2])
    peaks = []                # (center_px, lambda_nm)
    for row in wb["Sheet2"].iter_rows(values_only=True):
        if isinstance(row[0], (int, float)) and isinstance(row[1], (int, float)):
            if row[1] > 100.0:                    # wavelength row, not constants
                peaks.append((float(row[0]), float(row[1])))
    exposure = None
    for row in wb["Sheet1"].iter_rows(max_row=21, values_only=True):
        text = str(row[2])
        if "frame_exposure" in text:
            exposure = float(text.split("=")[1])
    return (np.asarray(px, float), np.asarray(cnt, float),
            (a0, a1, a2), peaks, exposure)


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    px, cnt, (a0, a1, a2), peaks, exposure = load_workbook_data(xlsx)
    lam = a0 + a1 * px + a2 * px ** 2

    plt.rcParams.update({"font.size": 15})
    fig, ax = plt.subplots(figsize=(2381 / 120, 880 / 120))
    ax.plot(lam, cnt, lw=0.8, color="tab:blue")
    ax.set_xlabel("Wavelength (nm)")
    ax.set_ylabel("Counts (ADU)")
    ax.set_xlim(350, 750)
    ax.set_ylim(0, 70000)
    ax.set_axisbelow(True)
    ax.grid(alpha=0.35)

    for center_px, line_nm in peaks:
        window = np.abs(px - center_px) < 6
        height = cnt[window].max()
        ax.annotate(f"{line_nm:.1f}",
                    xy=(a0 + a1 * center_px + a2 * center_px ** 2,
                        height + 1500),
                    ha="center", fontsize=14)

    # Inset: Gaussian fit to the isolated Hg 404.7 nm line.
    hg_px = min(peaks, key=lambda p: abs(p[1] - HG_LINE_NM))[0]
    m = np.abs(px - hg_px) <= FIT_HALF_PX
    popt, _ = curve_fit(gauss, px[m], cnt[m],
                        p0=[cnt[m].max() - np.median(cnt), hg_px, 2.0,
                            np.median(cnt)],
                        maxfev=40000)
    amp, mu, sigma, off = popt
    sigma = abs(sigma)
    disp = abs(a1 + 2 * a2 * mu)                  # nm per pixel at the line
    fwhm_nm = 2.3548 * sigma * disp

    axins = ax.inset_axes(INSET_BOX)
    lam_c = a0 + a1 * mu + a2 * mu ** 2
    show = (lam >= INSET_XLIM[0]) & (lam <= INSET_XLIM[1])
    axins.plot(lam[show], cnt[show], "o", ms=4, color="tab:blue", alpha=0.7)
    px_fine = np.linspace(px[show].min(), px[show].max(), 400)
    lam_fine = a0 + a1 * px_fine + a2 * px_fine ** 2
    axins.plot(lam_fine, gauss(px_fine, *popt), color="tab:red", lw=1.8)
    half_y = off + 0.5 * amp
    half_dx = sigma * np.sqrt(2 * np.log(2)) * disp
    axins.plot([lam_c - half_dx, lam_c + half_dx], [half_y, half_y],
               "--", color="tab:red", lw=1.8)
    axins.set_title(f"Hg 404.7 nm\nFWHM = {fwhm_nm:.2f} nm", fontsize=13)
    axins.set_xlabel("nm", fontsize=12)
    axins.set_ylabel("ADU", fontsize=12)
    axins.set_xlim(*INSET_XLIM)
    axins.tick_params(labelsize=11)
    axins.set_axisbelow(True)
    axins.grid(alpha=0.35)

    fig.subplots_adjust(**MAIN_AXES_RECT)
    fig.savefig(OUT_PNG, dpi=120)
    exp_ms = f"{exposure * 1e3:.0f} ms" if exposure else "unknown"
    print(f"Wrote {OUT_PNG}")
    print(f"exposure {exp_ms}; Hg 404.7 nm FWHM = {fwhm_nm:.4f} nm "
          f"(center px {mu:.3f}, sigma {sigma:.3f} px, {disp:.5f} nm/px)")


if __name__ == "__main__":
    main()
