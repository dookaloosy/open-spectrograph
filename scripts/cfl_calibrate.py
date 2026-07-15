"""Recalibrate the wavelength axis from a CFL spectrum capture.

Thin CLI over `controller.calibrate` / `controller.protocol`.  Accepts:

  *.tcd1304  — native capture (TCD1304DeviceMP format).  Frames are
               averaged (stale clipped frames dropped), index lines
               come from data/cfl_lines.toml (wavelengths only; the
               pattern locator finds their positions), and the
               ready-to-paste `store coefficients` controller command
               is printed.

  *.xlsx     — capture workbook (Sheet1 pixel/counts, Sheet2 line
               list).  Constants and fitted centers are written back
               into the workbook by editing the sheet XML inside the
               xlsx (openpyxl round-trips drop charts), and Excel is
               set to recalculate the wavelength column on load.

Neither path writes to data/cfl_lines.toml — it is a static list of
reference wavelengths.

Usage:
    python3 scripts/cfl_calibrate.py [capture.xlsx | capture.tcd1304]
"""

import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from controller import calibrate as cal
from controller import protocol

DEFAULT_XLSX = (cal.REPO / "output" / "export_czerny_baseline_v0_design"
                / "cfl_spectrum.xlsx")


def load_xlsx(path: Path):
    """Sheet1 pixel/counts + exposure, Sheet2 seeds."""
    import numpy as np
    wb = openpyxl.load_workbook(path, data_only=True)
    px, cnt = [], []
    exposure = None
    for row in wb["Sheet1"].iter_rows(values_only=True):
        if isinstance(row[0], (int, float)) and isinstance(row[2], (int, float)):
            px.append(row[0])
            cnt.append(row[2])
        elif "frame_exposure" in str(row[2]):
            exposure = float(str(row[2]).split("=")[1])
    seeds = []
    for row in wb["Sheet2"].iter_rows(values_only=True):
        if (isinstance(row[0], (int, float)) and isinstance(row[1], (int, float))
                and row[1] > 100.0):
            seeds.append((float(row[0]), float(row[1])))
    return np.asarray(px, float), np.asarray(cnt, float), seeds, exposure


def write_workbook(path: Path, centers_by_row, a0, a1, a2):
    """Chart-safe write-back: edit sheet XML inside the xlsx package."""
    z = zipfile.ZipFile(path)
    s1 = z.read("xl/worksheets/sheet1.xml").decode()
    s2 = z.read("xl/worksheets/sheet2.xml").decode()
    wbxml = z.read("xl/workbook.xml").decode()

    def setcell(xml, ref, val):
        pat = r'(<c r="%s"[^>]*><v>)[^<]*(</v>)' % ref
        new, n = re.subn(pat, r"\g<1>%s\g<2>" % val, xml)
        if n != 1:
            raise RuntimeError(f"cell {ref}: expected 1 match, got {n}")
        return new

    consts = [f"{a0:.6f}", f"{a1:.8f}", f"{a2:.6E}"]
    for ref, val in zip(("B18", "B19", "B20"), consts):
        s1 = setcell(s1, ref, val)
    for i, center in enumerate(centers_by_row, start=1):
        s2 = setcell(s2, f"A{i}", f"{center:.3f}")
    for ref, val in zip(("A8", "B8", "A9"), consts):
        s2 = setcell(s2, ref, val)

    if "fullCalcOnLoad" in wbxml:
        wbxml = re.sub(r'fullCalcOnLoad="[^"]*"', 'fullCalcOnLoad="1"', wbxml)
    elif "<calcPr" in wbxml:
        wbxml = wbxml.replace("<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
    else:
        wbxml = wbxml.replace(
            "</workbook>", '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')

    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for info in z.infolist():
            data = z.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                data = s1.encode()
            elif info.filename == "xl/worksheets/sheet2.xml":
                data = s2.encode()
            elif info.filename == "xl/workbook.xml":
                data = wbxml.encode()
            out.writestr(info, data)
    z.close()
    shutil.move(tmp, str(path))


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    seeds = None
    if src.suffix == ".tcd1304":
        capture = protocol.parse_capture(src)
        px, cnt, n_used, n_dropped = protocol.mean_spectrum(capture)
        if n_dropped:
            print(f"dropped {n_dropped} mostly-clipped frame(s)")
        lines, dispersion = cal.load_lines()
        print(f"{src.name}: {n_used} frame(s) averaged, "
              f"index lines from {cal.SEEDS_TOML.name}")
        exposure = capture.exposure_s
    else:
        px, cnt, seeds, exposure = load_xlsx(src)
        # wavelengths from the workbook's line list; nominal dispersion
        # derived from its recorded positions
        lines = sorted(lam for _, lam in seeds)
        (p_hi, l_hi), (p_lo, l_lo) = seeds[0], seeds[-1]
        dispersion = (l_lo - l_hi) / (p_lo - p_hi)

    result = cal.calibrate(px, cnt, lines, dispersion)

    exp_ms = f"{exposure * 1e3:.0f} ms" if exposure else "unknown"
    print(f"{src.name}: exposure {exp_ms}, "
          f"baseline {result.baseline_adu:.0f} ADU, "
          f"{result.n_clipped_px} clipped px\n")
    print(cal.report(result))

    if src.suffix == ".tcd1304":
        print(f"\ncontroller command:")
        print(result.store_command())
    else:
        seed_order = [lam for _, lam in seeds]
        center_of = {f.wavelength_nm: f.center_px for f in result.lines}
        write_workbook(src, [center_of[lam] for lam in seed_order],
                       result.a0, result.a1, result.a2)
        print(f"\nWrote constants and centers back into {src}")


if __name__ == "__main__":
    main()
